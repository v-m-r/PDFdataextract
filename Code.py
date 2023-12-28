from openpyxl import *
from datetime import datetime
from PyPDF2 import PdfMerger
from PyPDF2 import PdfReader
import os
import re
class pdf:
    path=r'Folder/you/have/pdf'
    common_name='.pdf'
    def __init__(self):
        print('PDF Starts')


    def find_pdf(self):
        print("finding pdfs")
        files = []
        for root, dirs, filenames in os.walk(self.path):
            for filename in filenames:
                if self.common_name in filename:
                    files.append(os.path.join(root, filename))
        print("found pdfs")
        return files


    def merge(self,file):
        print("merger started")

        self.pdfm = PdfMerger()
        for i in file:
            self.pdfm.append(i)
        self.pdfm.write("MergedPDF "+str(datetime.now().strftime("%d-%m-%Y"))+".pdf")
        self.pdfm.close()
        print("merger Ends") 
    def __del__(self):
        print("PDF Ends")

class excel:
    pdf = "MergedPDF "+str(datetime.now().strftime("%d-%m-%Y"))+".pdf"
    def __init__(self):
        print("Excel Part")
    
    def Data_entry(self):
        print("Data Entry start")
        self.reader= PdfReader(self.pdf)
        self.count=len(self.reader.pages)
        self.wb = load_workbook('Data.xlsx')
        self.ws = self.wb.active
        for i in range(0,self.count):
            self.page=self.reader.pages[i]
            self.text = self.page.extract_text()
            self.line = self.text.split('\n')
            if(self.line[23]=='Address'):
                self.adr = self.line[24]
            else:
                self.adr=self.line[23]
            self.apno=str(re.findall('Application File Number:\n[A-Z0-9_]*',self.text))
            self.sn=str(re.findall('Surname:[A-Z\s]*',self.text))
            self.n=str(re.findall('Given Name:[A-Z\s]*',self.text))
            self.fn=str(re.findall('Father\'s/LG\'s Name:[A-Z\s]*',self.text))
            self.ws['A'+str(i+2)]=self.apno[28:-2]
            self.ws['B'+str(i+2)]=self.n[13:-5]+self.sn[10:-5]
            self.ws['C'+str(i+2)]=self.adr
            self.ws['D'+str(i+2)]=self.fn[21:-5]
        self.wb.save('Data.xlsx')
        print("Data Entry End")
    def __del__(self):
        print("Excel end")


if __name__=="__main__":
    count =0 
    p=pdf()
    print()
    xl=excel()
    print()
    file=p.find_pdf()
    print()
    p.merge(file)
    print()
    print()
    xl.Data_entry()
    print()
